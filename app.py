import openpyxl as xl
from openpyxl.chart import BarChart, Reference


wb = xl.load_workbook('transactions.xlsx')
sheet = wb.active


sheet.cell(row=1, column=4).value = "Corrected Price"

for row in range(2, sheet.max_row + 1):
    price_cell = sheet.cell(row, 3)
    if isinstance(price_cell.value, (int, float)):
        corrected_price = price_cell.value * 0.9
        sheet.cell(row, 4).value = corrected_price


data = Reference(sheet, min_col=4, min_row=1, max_row=sheet.max_row)
chart = BarChart()
chart.add_data(data, titles_from_data=True)
chart.title = "Corrected Prices"
chart.x_axis.title = "Transaction #"
chart.y_axis.title = "Price ($)"


sheet.add_chart(chart, "A8")


wb.save('transactions1.xlsx')




